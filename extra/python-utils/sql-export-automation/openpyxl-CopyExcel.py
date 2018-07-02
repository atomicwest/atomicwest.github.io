#====================================
#Author: Jesson Go
#Date: 20180701
#copy excel sheet with formatting
#need to know max columns used

#====================================

import openpyxl as ox
from copy import copy
import inspect
import datetime as dt



def excelCols(n):
    output = ''
    numLetters = int(n/26)+1
    # front = chr(numLetters+65)
    if n < 26:
        output+=chr(n + 65)
    else:
        firstletterNum = int(n/26)
        output+=chr(firstletterNum + 64)
        
        padding = n-(26*int(n/26))
        output+=chr(padding+65)
    return output


def npad(n, padLength=2):
    num =  str(n) 
    if len(num) < padLength:
        num = '0'*(padLength-len(num)) + num
    return num



def copySheetAsBook(sourcePath, savePath, sheetTitle):
    sourceBook = ox.load_workbook(sourcePath)
    sourceSheet = sourceBook.worksheets[0]
    destBook = ox.Workbook()
    
##    currentSheet = destBook.create_sheet(title='MappingDocs')
    currentSheet = destBook.active
    currentSheet.title = sheetTitle
    
    rangeString = 'A{}:E{}'.format(sourceSheet.min_row, sourceSheet.max_row)

    for row in sourceSheet[rangeString]:
        for cell in row:
            cellstyle = ''
            if cell.has_style:
                cellstyle = cell.font
##                print(cell.font.color)
            #------------write stage----------------
            cellNum = cell.column+str(cell.row)
            currentSheet[cellNum] = cell.value
            currentSheet[cellNum].font = copy(cell.font)
##            print('%s : %s | %s' % ( cellNum, cell.value, cellstyle ))
    
    destBook.save(savePath)



def copyExcel(sourcePath, savePath):
    sourceBook = ox.load_workbook(sourcePath)
    destBook = ox.Workbook()
    for i in range(len(sourceBook.worksheets)):
        currentSourceSheet = sourceBook.worksheets[i]
        currentDestSheet = destBook.create_sheet(title=currentSourceSheet.title)
        rangeString = 'A{}:{}{}'.format(currentSourceSheet.min_row, excelCols(currentSourceSheet.max_column-1), currentSourceSheet.max_row)

##        print(currentDestSheet.row_dimensions)
##        print(currentDestSheet.column_dimensions)
        
        for row in currentSourceSheet[rangeString]:
            for cell in row:
##                print(currentDestSheet.row_dimensions[cell.row].height)
##                print(currentDestSheet.column_dimensions[cell.column].width)
##                print('column width: ' + str(currentSourceSheet.column_dimensions[cell.column].width) )
##                print('row height: ' + str(currentSourceSheet.row_dimensions[cell.row].height))     #dont need to convert cell.row to string
##                print(cell.coordinate)
##                print(inspect.getmembers(cell))
##                cellNum = cell.column+str(cell.row)
                currentDestSheet[cell.coordinate] = cell.value
                currentDestSheet[cell.coordinate].font = copy(cell.font)
                if type(currentSourceSheet[cell.coordinate].comment) is ox.comments.Comment:
##                    print(currentSourceSheet[cell.coordinate].comment)
                    currentDestSheet[cell.coordinate].comment = copy(currentSourceSheet[cell.coordinate].comment)

                
                if currentSourceSheet.row_dimensions[cell.row].height is not None:
                    currentDestSheet.row_dimensions[cell.row].height = copy(currentSourceSheet.row_dimensions[cell.row].height)
                if currentSourceSheet.column_dimensions[cell.column].width is not None:
                    currentDestSheet.column_dimensions[cell.column].width = copy(currentSourceSheet.column_dimensions[cell.column].width)
                
    destBook.remove(destBook['Sheet'])   #don't know how else to get rid of the libary's default worksheet
    destBook.save(savePath)



d = dt.datetime.now()
datestamp = '_%s%s%s_%s%s' % (str(d.year), str(npad(d.month)), str(npad(d.day)), str(npad(d.hour)), str(npad(d.minute)) )
sourcePath = 'DataMappingTemplates.xlsx'
savePath = 'StyleSaveOPX_Sheets_%s.xlsx' % datestamp
sheetTitle = 'MappingDocs'
##copySheetAsBook(sourcePath, savePath, sheetTitle)
copyExcel(sourcePath, savePath)
