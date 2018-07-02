#====================================
#Author: Jesson Go
#Date: 20180701
#copy excel sheet with formatting
#need to know max columns used

#====================================

import openpyxl as ox
from copy import copy
import inspect
import datetime as dt
import pyodbc
import time as t

def excelCols(n):
    output = ''
    numLetters = int(n/26)+1
    # front = chr(numLetters+65)
    if n < 26:
        output+=chr(n + 65)
    else:
        firstletterNum = int(n/26)
        output+=chr(firstletterNum + 64)
        
        padding = n-(26*int(n/26))
        output+=chr(padding+65)
    return output


def npad(n, padLength=2):
    num =  str(n) 
    if len(num) < padLength:
        num = '0'*(padLength-len(num)) + num
    return num



#-----------this function is built assuming that a template is first copied to an output file-----------------------
#-----------then a sql export is saved to the output file-----------------------------------------------------------
def massExport(sourcePath, savePath, tableName, templateName):
    destBook = ox.Workbook()

    #-------load templates first----------------
    sourceBook = ox.load_workbook(sourcePath)
    #uncomment below to copy all of the templates
##    for i in range(len(sourceBook.worksheets)):
##        currentSourceSheet = sourceBook.worksheets[i]
##        currentDestSheet = destBook.create_sheet(title=currentSourceSheet.title)
##        rangeString = 'A{}:{}{}'.format(currentSourceSheet.min_row, excelCols(currentSourceSheet.max_column-1), currentSourceSheet.max_row)
##        
##        for row in currentSourceSheet[rangeString]:
##            for cell in row:
##                currentDestSheet[cell.coordinate] = cell.value
##                currentDestSheet[cell.coordinate].font = copy(cell.font)
##                if type(currentSourceSheet[cell.coordinate].comment) is ox.comments.Comment:
##                    currentDestSheet[cell.coordinate].comment = copy(currentSourceSheet[cell.coordinate].comment)
##
##                
##                if currentSourceSheet.row_dimensions[cell.row].height is not None:
##                    currentDestSheet.row_dimensions[cell.row].height = copy(currentSourceSheet.row_dimensions[cell.row].height)
##                if currentSourceSheet.column_dimensions[cell.column].width is not None:
##                    currentDestSheet.column_dimensions[cell.column].width = copy(currentSourceSheet.column_dimensions[cell.column].width)

    currentSourceSheet = sourceBook[templateName]
    currentDestSheet = destBook.create_sheet(title=currentSourceSheet.title)
    rangeString = 'A{}:{}{}'.format(currentSourceSheet.min_row, excelCols(currentSourceSheet.max_column-1), currentSourceSheet.max_row)
    
    for row in currentSourceSheet[rangeString]:
        for cell in row:
            currentDestSheet[cell.coordinate] = cell.value
            currentDestSheet[cell.coordinate].font = copy(cell.font)
            if type(currentSourceSheet[cell.coordinate].comment) is ox.comments.Comment:
                currentDestSheet[cell.coordinate].comment = copy(currentSourceSheet[cell.coordinate].comment)

            
            if currentSourceSheet.row_dimensions[cell.row].height is not None:
                currentDestSheet.row_dimensions[cell.row].height = copy(currentSourceSheet.row_dimensions[cell.row].height)
            if currentSourceSheet.column_dimensions[cell.column].width is not None:
                currentDestSheet.column_dimensions[cell.column].width = copy(currentSourceSheet.column_dimensions[cell.column].width)

    #-------add sql exports----------------

    exportSheet = destBook.create_sheet(title='%sExports'%tableName)
    #create SSMS connection
    server = ''
    database = '' 
    username = '' 
    password = '' 
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};Trusted_connection=yes;SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()

    getcols = """
              SELECT column_name from INFORMATION_SCHEMA.columns
              WHERE table_name LIKE '%%%s%%'
              """ % tableName

    #- - - - - - add headers first - - - - - -
    header = []
    cursor.execute(getcols)
    headerrow = cursor.fetchone()
    while headerrow:
        header.append(headerrow[0])
        headerrow = cursor.fetchone()
    
    maxCol = excelCols(len(header)-1)
    bold = ox.styles.Font(bold=True)
##    print(header)
    for i,h in enumerate(header):
        coord = excelCols(i) + '1'
        exportSheet[coord] = h
        exportSheet[coord].font = bold


    #- - - - - - add data - - - - - -
    query = """
            SELECT * FROM %s
            """ % (tableName)

    query2 = """
            SELECT * FROM %s
            WHERE colB LIKE '%%something%%'
            OR colB LIKE '%%%s%%'
            """ % (tableName, 'nothing')

    out = []

    cursor.execute(query) 
    row = cursor.fetchone() 
    while row:
        #parse into cells in the new sheet
        out.append(row)
        row = cursor.fetchone()

    for i,row in enumerate(out):
        for j,data in enumerate(row):
            coord = excelCols(j) + str(i+2)
            exportSheet[coord] = data
    
    #-------add extra tab at the end----------------
    
    feedbackSheet = destBook.create_sheet(title='Feedback')
    destBook.remove(destBook['Sheet'])   #don't know how else to get rid of the libary's default worksheet
    
    #-------complete export----------------
    destBook.save(savePath)


d = dt.datetime.now()
datestamp = '_%s%s%s_%s%s' % (str(d.year), str(npad(d.month)), str(npad(d.day)), str(npad(d.hour)), str(npad(d.minute)) )
sourcePath = 'Templates.xlsx'
savePath = 'StyleSaveOPX_Sheets_%s.xlsx' % (datestamp)
tableName = 'myTable'
templateName = 'templateTab'

tStart = t.time()
massExport(sourcePath, savePath,tableName, templateName)
tStop = t.time()
print("Operation completed in %s minutes" % (round((tStop-tStart)/60,4) ) )
